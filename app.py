import json
import gzip
import hmac
import os
import re
import shutil
import tempfile
import threading
from datetime import datetime
from functools import wraps
from pathlib import Path
from urllib.parse import urlparse

from dotenv import load_dotenv
load_dotenv()

import requests
from flask import Flask, Response, jsonify, redirect, render_template, request, send_file, session, url_for
from openpyxl import load_workbook
from juyuan_update import config as juyuan_config
from juyuan_update.tasks import get_status as get_update_status, start_update
from juyuan_update.unified_excel import (
    get_bond_picker_bonds,
    import_unified_excel,
    load_bond_picker_yields_cache,
    load_bond_static,
    load_update_settings,
    save_update_settings,
)
from primary_market_pricing.app import pricing_bp
from internal_knowledge_base import bp as internal_knowledge_base_bp
import institution_flow_config
import institution_flow_data

# 运态数据路径统一由 paths.py 管理（PORTAL_DATA_ROOT 环境变量定位）
from paths import (
    BASE_DIR,
    DATA_DIR,
    BOND_DIR,
    STRATEGY_DIR,
    SPREAD_DIR,
    INDUSTRY_DIR,
    STD_DEV_DIR,
    CONFIG_DIR,
    UPLOADS_DIR,
    PRIMARY_PRICING_CACHE,
)

BOND_EXCEL = BOND_DIR / "数据.xlsx"
STRATEGY_HTML = STRATEGY_DIR / "信用债策略仪表盘.html"
SPREAD_JS = SPREAD_DIR / "spread_data.js"
INDUSTRY_HTML = INDUSTRY_DIR / "行业景气度跟踪.html"
STD_DEV_HTML = STD_DEV_DIR / "index.html"
STD_DEV_JS = STD_DEV_DIR / "data" / "spread_data.js"
INSTITUTION_FLOW_CACHE = institution_flow_config.RATE_CURVES_CACHE
MAPPING_FILE = CONFIG_DIR / "映射表.xlsx"
SHEET_NAME = "万得"

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-change-before-deploy")
app.register_blueprint(pricing_bp, url_prefix="/primary-market-pricing")
app.register_blueprint(internal_knowledge_base_bp, url_prefix="/internal-knowledge-base")

BONDS_CACHE = []
DATA_TIMESTAMP = "尚未加载"
INSTITUTION_FLOW_UPSTREAM = "http://43.137.12.140:8000/bondflow/api"
INSTITUTION_FLOW_UPSTREAM_PAGE = "http://43.137.12.140:8000/jgxw/"
INSTITUTION_FLOW_TIMEOUT = 120
INSTITUTION_FLOW_OPTIONS_TTL = 300
INSTITUTION_FLOW_QUERY_TTL = int(os.environ.get("INSTITUTION_FLOW_QUERY_TTL", "30"))
INSTITUTION_FLOW_CACHE_MAX = int(os.environ.get("INSTITUTION_FLOW_CACHE_MAX", "128"))
_institution_flow_options = {"timestamp": 0.0, "payload": None}
_institution_flow_cache_lock = threading.Lock()
_institution_flow_cache = {}
_institution_flow_http = requests.Session()
_institution_flow_http.headers.update({
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) credit-tools-portal/1.0",
    "Referer": INSTITUTION_FLOW_UPSTREAM_PAGE,
})

COMPRESS_MIN_SIZE = int(os.environ.get("COMPRESS_MIN_SIZE", "1024"))
COMPRESS_LEVEL = max(1, min(9, int(os.environ.get("COMPRESS_LEVEL", "5"))))
_compress_cache_lock = threading.Lock()
_compress_cache = {}
_compressible_mimetypes = {
    "application/javascript",
    "application/json",
    "application/xml",
    "image/svg+xml",
    "text/css",
    "text/html",
    "text/javascript",
    "text/plain",
    "text/xml",
}


def site_password():
    return os.environ.get("SITE_PASSWORD", "Abcd123%")


def admin_password():
    return os.environ.get("ADMIN_PASSWORD", "123456")


def safe_next_path(value, fallback=None):
    fallback = fallback or url_for("home")
    if not value:
        return fallback
    parsed = urlparse(value)
    return value if not parsed.scheme and not parsed.netloc and value.startswith("/") else fallback


def login_required(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        if not session.get("authenticated"):
            return redirect(url_for("login", next=request.path))
        return func(*args, **kwargs)
    return wrapper


def admin_required(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        if not session.get("admin_authenticated"):
            return redirect(url_for("admin_login", next=request.path))
        return func(*args, **kwargs)
    return wrapper


@app.before_request
def protect_primary_market_pricing():
    if request.blueprint == pricing_bp.name and not session.get("authenticated"):
        return redirect(url_for("login", next=request.path))



def format_date_only(value):
    if not value:
        return ""
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text[:19], fmt).strftime("%Y - %m - %d")
        except ValueError:
            pass
    if len(text) >= 10:
        head = text[:10].replace("/", "-")
        parts = head.split("-")
        if len(parts) == 3 and all(p.isdigit() for p in parts):
            return f"{int(parts[0]):04d} - {int(parts[1]):02d} - {int(parts[2]):02d}"
    return text

def file_updated(path: Path):
    if not path.exists():
        return "文件不存在"
    return datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y - %m - %d")


def file_size(path: Path):
    if not path.exists():
        return "-"
    size = path.stat().st_size
    if size >= 1024 * 1024:
        return f"{size / 1024 / 1024:.1f} MB"
    if size >= 1024:
        return f"{size / 1024:.1f} KB"
    return f"{size} B"


def file_updated_time(path: Path):
    if not path.exists():
        return "文件不存在"
    return datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y - %m - %d %H:%M")


def read_excel(path: Path):
    static_bonds = get_bond_picker_bonds()
    if static_bonds:
        yield_cache = load_bond_picker_yields_cache()
        yields = yield_cache.get("yields") or {}
        bonds = []
        for bond in static_bonds:
            code = str(bond.get("code") or "").strip().upper()
            symbol = code.split(".", 1)[0]
            ytm = yields.get(code)
            if ytm is None:
                ytm = yields.get(symbol)
            if ytm is None:
                continue
            bonds.append([
                bond.get("code") or "",
                bond.get("name") or "",
                round(float(bond.get("term") or 0), 4),
                bond.get("implied_rating") or "",
                bond.get("issuer") or "",
                round(float(ytm), 4),
                bond.get("entity") or "",
                bond.get("ct") or "",
                bond.get("sub") or "",
                bond.get("tech") or "",
                bond.get("internal_rating") or "",
            ])
        data_date = yield_cache.get("trade_date") or load_bond_static().get("generated_at") or ""
        return bonds, data_date

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]

    n1_value = ws.cell(row=1, column=14).value
    data_date = str(n1_value).strip() if n1_value else ""
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    bonds = []
    for row in rows:
        if not row or len(row) < 11:
            continue
        code = str(row[0] or "").strip()
        name = str(row[1] or "").strip()
        term = row[2]
        rating = str(row[3] or "").strip()
        issuer = str(row[4] or "").strip()
        ytm = row[5]
        entity = str(row[6] or "").strip()
        ct = str(row[7] or "").strip()
        sub = str(row[8] or "").strip()
        tech = str(row[9] or "").strip()
        ir = str(row[10] or "").strip()
        if not code or not name or term is None or ytm is None:
            continue
        try:
            term = round(float(term), 4)
            ytm = round(float(ytm), 4)
        except (TypeError, ValueError):
            continue
        bonds.append([code, name, term, rating, issuer, ytm, entity, ct, sub, tech, ir])
    return bonds, data_date


def load_bond_data():
    global BONDS_CACHE, DATA_TIMESTAMP
    if not BOND_EXCEL.exists() and not juyuan_config.BOND_STATIC_JSON.exists():
        BONDS_CACHE = []
        DATA_TIMESTAMP = "数据文件未找到"
        return
    try:
        bonds, data_date = read_excel(BOND_EXCEL)
        BONDS_CACHE = bonds
        DATA_TIMESTAMP = format_date_only(data_date) or file_updated(BOND_EXCEL)
    except Exception as exc:
        BONDS_CACHE = []
        DATA_TIMESTAMP = f"读取失败: {exc}"


def status_info():
    load_bond_data()
    static_payload = load_bond_static()
    unified_updated = file_updated_time(juyuan_config.UNIFIED_EXCEL)
    settings = load_update_settings()
    return {
        "bond_picker": {
            "updated": DATA_TIMESTAMP,
            "total": f"{len(BONDS_CACHE):,}",
        },
        "strategy_dashboard": {
            "updated": file_updated(STRATEGY_HTML),
            "size": file_size(STRATEGY_HTML),
        },
        "spread_monitor": {
            "updated": file_updated(SPREAD_JS),
            "size": file_size(SPREAD_JS),
            "bond_list_updated": file_updated_time(juyuan_config.PROJECT2_BOND_EXCEL),
            "bond_list_size": file_size(juyuan_config.PROJECT2_BOND_EXCEL),
        },
        "industry_prosperity": {
            "updated": file_updated(INDUSTRY_HTML),
            "size": file_size(INDUSTRY_HTML),
        },
        "credit_std_dev": {
            "updated": file_updated(STD_DEV_HTML),
            "size": file_size(STD_DEV_HTML),
            "data_updated": file_updated(STD_DEV_JS),
            "data_size": file_size(STD_DEV_JS),
        },
        "primary_market_pricing": {
            "updated": file_updated_time(PRIMARY_PRICING_CACHE),
            "size": file_size(PRIMARY_PRICING_CACHE),
        },
        "institution_flow": {
            "updated": file_updated_time(INSTITUTION_FLOW_CACHE),
            "size": file_size(INSTITUTION_FLOW_CACHE),
        },
        "unified_excel": {
            "updated": unified_updated,
            "size": file_size(juyuan_config.UNIFIED_EXCEL),
            "total_bonds": static_payload.get("total_bonds", 0),
            "source_generated": static_payload.get("generated_at") or "-",
            "fund_updated": file_updated_time(juyuan_config.STRATEGY_FUND_PRICES_FROZEN),
        },
        "settings": settings,
    }


def portal_nav(active_endpoint: str):
    links = [
        ("home", "首页"),
        ("bond_picker", "择券工具"),
        ("strategy_dashboard", "策略仪表盘"),
        ("spread_monitor", "利差监控"),
        ("primary_market_pricing.index", "一级发行研究"),
        ("industry_prosperity", "行业景气度"),
        ("credit_std_dev", "两倍标准差"),
        ("institution_flow", "机构行为监测"),
        ("internal_knowledge_base.index", "内部知识库"),
        ("admin", "后台上传"),
    ]
    items = ['<style id="portal-nav-style">']
    items.append(
        ".portal-nav{position:sticky;top:0;z-index:10000;background:#fff;color:#1e293b;"
        "height:44px;display:flex;align-items:center;gap:6px;padding:0 16px;"
        "border-bottom:1px solid #e2e8f0;font-family:-apple-system,BlinkMacSystemFont,"
        '"Segoe UI","Microsoft YaHei",sans-serif;box-sizing:border-box}'
        ".portal-nav *{box-sizing:border-box}"
        ".portal-nav .brand{font-weight:800;color:#2563eb;font-size:15px;margin-right:16px;letter-spacing:.5px}"
        ".portal-nav a{color:#334155;text-decoration:none;font-size:12.5px;font-weight:600;"
        "padding:0 14px;height:30px;display:inline-flex;align-items:center;border-radius:4px;transition:all .2s}"
        ".portal-nav a:hover{color:#2563eb;background:#eff6ff}"
        ".portal-nav a.active{color:#2563eb;background:#eff6ff;font-weight:700}"
        ".portal-nav a.management-link{margin-left:8px;border-left:1px solid #cbd5e1;border-radius:0 4px 4px 0;padding-left:18px;color:#475569}"
    )
    items.append("</style>")
    items.append('<!-- CREDIT_TOOLS_PORTAL_NAV -->')
    items.append('<div class="portal-nav" data-portal-nav="credit-tools">')
    items.append('<div class="brand">内部研究平台</div>')
    for endpoint, label in links:
        classes = []
        if endpoint == active_endpoint:
            classes.append("active")
        if endpoint == "internal_knowledge_base.index":
            classes.append("management-link")
        class_attr = f' class="{" ".join(classes)}"' if classes else ""
        items.append(f'<a href="{url_for(endpoint)}"{class_attr}>{label}</a>')
    items.append("</div>")
    return "".join(items)


def find_matching_div_end(html: str, start: int):
    depth = 0
    for match in re.finditer(r"</?div\b[^>]*>", html[start:], flags=re.I):
        token = match.group(0)
        if token.startswith("</"):
            depth -= 1
            if depth == 0:
                return start + match.end()
        else:
            depth += 1
    return -1


def remove_portal_nav_divs(html: str):
    pattern = re.compile(
        r"<div\b(?=[^>]*(?:data-portal-nav=|class=[\"'][^\"']*\bportal-nav\b))[^>]*>",
        flags=re.I,
    )
    result = html
    search_from = 0
    while True:
        match = pattern.search(result, search_from)
        if not match:
            return result
        end = find_matching_div_end(result, match.start())
        if end == -1:
            search_from = match.end()
            continue
        result = result[:match.start()] + result[end:]
        search_from = match.start()


def remove_portal_nav_css(html: str):
    cleaned = re.sub(r"<style\b[^>]*id=[\"']portal-nav-style[\"'][^>]*>.*?</style>\s*", "", html, flags=re.I | re.S)
    cleaned = re.sub(r"\s*\.portal-nav[^{}]*\{[^{}]*\}", "", cleaned, flags=re.I | re.S)
    return cleaned


def insert_portal_nav(html: str, nav: str):
    head_match = re.search(r"</head\s*>", html, flags=re.I)
    body_search_start = head_match.end() if head_match else 0
    body_match = re.search(r"<body\b[^>]*>", html[body_search_start:], flags=re.I)
    if body_match:
        insert_at = body_search_start + body_match.end()
        return html[:insert_at] + nav + html[insert_at:]
    if head_match:
        return html[:head_match.end()] + nav + html[head_match.end():]
    return nav + html


def inject_portal_nav(html: str, active_endpoint: str):
    cleaned = re.sub(r"<!-- CREDIT_TOOLS_PORTAL_NAV -->\s*", "", html, flags=re.I)
    cleaned = remove_portal_nav_divs(cleaned)
    cleaned = remove_portal_nav_css(cleaned)
    return insert_portal_nav(cleaned, portal_nav(active_endpoint))


def render_portal_template(template_name: str, active_endpoint: str, **context):
    html = render_template(template_name, **context)
    return Response(inject_portal_nav(html, active_endpoint), content_type="text/html; charset=utf-8")


# 注入导航栏后的 HTML 缓存：仪表盘 HTML（如策略仪表盘 ~7MB）每次请求都全量读盘 +
# 全量正则注入极其昂贵，而文件仅在"一键更新"重写后才变化。按 mtime 缓存最终结果，
# 文件未变时直接复用，省去读盘与正则扫描。
_html_cache_lock = threading.Lock()
_html_cache = {}  # key -> {"mtime": float, "html": str}


def html_response(path: Path, active_endpoint: str, missing_message: str, replacements=None):
    if not path.exists():
        return Response(missing_message, status=404, content_type="text/plain; charset=utf-8")
    try:
        stat = path.stat()
        mtime = stat.st_mtime
        size = stat.st_size
    except OSError:
        mtime = None
        size = None

    cache_key = (
        str(path),
        active_endpoint,
        tuple(sorted(replacements.items())) if replacements else None,
    )

    html = None
    with _html_cache_lock:
        ent = _html_cache.get(cache_key)
        if ent and ent["mtime"] == mtime:
            html = ent["html"]

    if html is None:
        html = path.read_text(encoding="utf-8", errors="replace")
        if replacements:
            for old, new in replacements.items():
                html = html.replace(old, new)
        html = inject_portal_nav(html, active_endpoint)
        with _html_cache_lock:
            _html_cache[cache_key] = {"mtime": mtime, "html": html}

    # 浏览器缓存：1 分钟内不重复请求整页；ETag 命中时仅回 304，省去 7MB 传输。
    etag = f'"{active_endpoint}-{int(mtime or 0)}-{size or 0}"'
    resp = Response(html, content_type="text/html; charset=utf-8")
    resp.headers["ETag"] = etag
    resp.headers["Cache-Control"] = "private, max-age=60"
    resp.make_conditional(request)
    return resp


@app.after_request
def compress_text_response(response):
    """Gzip sizeable text responses when the reverse proxy did not compress them."""
    if (
        request.method == "HEAD"
        or response.status_code < 200
        or response.status_code in {204, 205, 304}
        or response.headers.get("Content-Encoding")
        or "gzip" not in request.accept_encodings
        or request.accept_encodings["gzip"] <= 0
        or response.mimetype not in _compressible_mimetypes
    ):
        return response

    content_length = response.content_length
    if content_length is not None and content_length < COMPRESS_MIN_SIZE:
        return response

    cache_key = None
    etag = response.headers.get("ETag")
    if etag and content_length:
        cache_key = (etag, content_length, response.mimetype, COMPRESS_LEVEL)
        with _compress_cache_lock:
            compressed = _compress_cache.get(cache_key)
        if compressed is not None:
            response.direct_passthrough = False
            response.set_data(compressed)
            response.headers["Content-Encoding"] = "gzip"
            response.vary.add("Accept-Encoding")
            if etag and not etag.startswith("W/"):
                response.headers["ETag"] = f"W/{etag}"
            return response

    response.direct_passthrough = False
    payload = response.get_data()
    if len(payload) < COMPRESS_MIN_SIZE:
        return response
    compressed = gzip.compress(payload, compresslevel=COMPRESS_LEVEL)
    if len(compressed) >= len(payload):
        return response

    if cache_key is not None:
        with _compress_cache_lock:
            if len(_compress_cache) >= 32:
                _compress_cache.pop(next(iter(_compress_cache)))
            _compress_cache[cache_key] = compressed
    response.set_data(compressed)
    response.headers["Content-Encoding"] = "gzip"
    response.vary.add("Accept-Encoding")
    if etag and not etag.startswith("W/"):
        response.headers["ETag"] = f"W/{etag}"
    return response


@app.after_request
def inject_primary_market_pricing_nav(response):
    if request.blueprint != pricing_bp.name or not response.content_type.startswith("text/html"):
        return response
    response.set_data(inject_portal_nav(response.get_data(as_text=True), "primary_market_pricing.index"))
    return response

BACKUP_RETENTION_COUNT = 3


def file_sha256(path: Path) -> str:
    import hashlib

    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def prune_upload_backups(backup_key: str, suffix: str) -> None:
    backups = sorted(
        UPLOADS_DIR.glob(f"{backup_key}_backup_*{suffix}"),
        key=lambda path: path.name,
        reverse=True,
    )
    for backup in backups[BACKUP_RETENTION_COUNT:]:
        backup.unlink(missing_ok=True)


def save_upload(file_storage, destination: Path, allowed_exts, backup_key: str) -> bool:
    """Replace an uploaded file atomically and retain recent distinct versions."""
    filename = file_storage.filename or ""
    ext = Path(filename).suffix.lower()
    if ext not in allowed_exts:
        raise ValueError(f"文件类型不正确，仅支持: {', '.join(sorted(allowed_exts))}")
    destination.parent.mkdir(parents=True, exist_ok=True)
    UPLOADS_DIR.mkdir(parents=True, exist_ok=True)

    fd, tmp_name = tempfile.mkstemp(prefix=f".{backup_key}-", suffix=".upload", dir=destination.parent)
    os.close(fd)
    staged = Path(tmp_name)
    try:
        file_storage.save(staged)
        if destination.exists() and file_sha256(staged) == file_sha256(destination):
            return False
        backup_created = False
        if destination.exists():
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
            backup_path = UPLOADS_DIR / f"{backup_key}_backup_{timestamp}{destination.suffix}"
            shutil.copy2(destination, backup_path)
            backup_created = True
        os.replace(staged, destination)
        if backup_created:
            prune_upload_backups(backup_key, destination.suffix)
        return True
    finally:
        staged.unlink(missing_ok=True)


@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("authenticated"):
        return redirect(url_for("home"))
    error = None
    if request.method == "POST":
        if request.form.get("password", "") == site_password():
            session["authenticated"] = True
            session["login_time"] = datetime.now().strftime("%Y-%m-%d %H:%M")
            return redirect(safe_next_path(request.args.get("next")))
        error = "访问密码错误"
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/admin/login", methods=["GET", "POST"])
@login_required
def admin_login():
    next_path = safe_next_path(request.args.get("next"), url_for("admin"))
    if session.get("admin_authenticated"):
        return redirect(next_path)
    error = None
    if request.method == "POST":
        submitted = request.form.get("password", "")
        if hmac.compare_digest(submitted, admin_password()):
            session["admin_authenticated"] = True
            session["admin_login_time"] = datetime.now().strftime("%Y-%m-%d %H:%M")
            return redirect(next_path)
        error = "后台密码错误"
    return render_template("admin_login.html", error=error)


@app.route("/admin/logout")
@login_required
def admin_logout():
    session.pop("admin_authenticated", None)
    session.pop("admin_login_time", None)
    return redirect(url_for("home"))


@app.route("/")
@login_required
def home():
    return render_portal_template("home.html", "home", status=status_info())


@app.route("/bond-picker")
@login_required
def bond_picker():
    load_bond_data()
    return render_portal_template(
        "bond_picker.html",
        "bond_picker",
        bond_data=json.dumps(BONDS_CACHE, ensure_ascii=False),
        timestamp=DATA_TIMESTAMP,
        total=len(BONDS_CACHE),
    )


@app.route("/strategy-dashboard")
@login_required
def strategy_dashboard():
    local_echarts = url_for("static", filename="vendor/echarts.min.js")
    return html_response(
        STRATEGY_HTML,
        "strategy_dashboard",
        "策略仪表盘 HTML 文件不存在，请先到后台上传。",
        replacements={
            "https://cdn.jsdelivr.net/npm/echarts@5/dist/echarts.min.js": local_echarts,
            "https://cdn.jsdelivr.net/npm/echarts@5.5.0/dist/echarts.min.js": local_echarts,
        },
    )


@app.route("/spread-monitor")
@login_required
def spread_monitor():
    return render_portal_template("spread_monitor.html", "spread_monitor")


@app.route("/data/spread_monitor/spread_data.js")
@login_required
def spread_data_js():
    if not SPREAD_JS.exists():
        return Response("var SPREAD_DATA = {data: [], total_bonds: 0, update_time: '数据文件未找到'};", mimetype="application/javascript; charset=utf-8")
    return send_file(SPREAD_JS, mimetype="application/javascript")


@app.route("/institution-flow")
@login_required
def institution_flow():
    return render_portal_template("institution_flow.html", "institution_flow")


def institution_flow_forward_args():
    return [
        (key, value)
        for key, values in request.args.lists()
        for value in values
    ]


@app.route("/institution-flow/bondflow/api/<path:subpath>")
@login_required
def institution_flow_proxy(subpath):
    import time

    is_options = subpath.rstrip("/") == "options" and not request.args
    if is_options:
        cached = _institution_flow_options
        if cached["payload"] is not None and time.time() - cached["timestamp"] < INSTITUTION_FLOW_OPTIONS_TTL:
            return jsonify(cached["payload"])
    cache_key = (subpath.rstrip("/"), tuple(sorted(institution_flow_forward_args())))
    if not is_options and INSTITUTION_FLOW_QUERY_TTL > 0:
        with _institution_flow_cache_lock:
            cached = _institution_flow_cache.get(cache_key)
        if cached and time.time() - cached["timestamp"] < INSTITUTION_FLOW_QUERY_TTL:
            return app.response_class(
                cached["content"], status=cached["status"], content_type=cached["content_type"],
            )
    try:
        response = _institution_flow_http.get(
            f"{INSTITUTION_FLOW_UPSTREAM}/{subpath}",
            params=institution_flow_forward_args(),
            timeout=INSTITUTION_FLOW_TIMEOUT,
        )
    except requests.RequestException as exc:
        return jsonify({"error": f"上游数据请求失败: {exc}"}), 502
    if is_options and response.ok:
        try:
            _institution_flow_options.update(timestamp=time.time(), payload=response.json())
        except ValueError:
            pass
    elif response.ok and INSTITUTION_FLOW_QUERY_TTL > 0:
        with _institution_flow_cache_lock:
            if len(_institution_flow_cache) >= INSTITUTION_FLOW_CACHE_MAX:
                _institution_flow_cache.pop(next(iter(_institution_flow_cache)))
            _institution_flow_cache[cache_key] = {
                "timestamp": time.time(),
                "content": response.content,
                "status": response.status_code,
                "content_type": response.headers.get("Content-Type", "application/json"),
            }
    return app.response_class(
        response.content,
        status=response.status_code,
        content_type=response.headers.get("Content-Type", "application/json"),
    )


@app.route("/institution-flow/api/overlay/meta")
@login_required
def institution_flow_overlay_meta():
    try:
        return jsonify(institution_flow_data.get_meta())
    except Exception as exc:
        return jsonify({"error": f"叠加曲线元数据加载失败: {exc}"}), 500


@app.route("/institution-flow/api/overlay/yield")
@login_required
def institution_flow_overlay_yield():
    curve = request.args.get("curve", "")
    tenor = request.args.get("tenor", "")
    if not curve or not tenor:
        return jsonify({"error": "缺少参数 curve / tenor"}), 400
    try:
        return jsonify(institution_flow_data.get_yield_series(curve, tenor))
    except Exception as exc:
        return jsonify({"error": f"收益率数据加载失败: {exc}"}), 500


@app.route("/institution-flow/api/overlay/spread")
@login_required
def institution_flow_overlay_spread():
    category = request.args.get("category", "")
    tenor = request.args.get("tenor", "")
    if not category or not tenor:
        return jsonify({"error": "缺少参数 category / tenor"}), 400
    try:
        return jsonify(institution_flow_data.get_spread_series(category, tenor))
    except Exception as exc:
        return jsonify({"error": f"利差数据加载失败: {exc}"}), 500


@app.route("/institution-flow/healthz")
@login_required
def institution_flow_healthz():
    return jsonify({"ok": True, "upstream": INSTITUTION_FLOW_UPSTREAM})



@app.route("/industry-prosperity")
@login_required
def industry_prosperity():
    return html_response(INDUSTRY_HTML, "industry_prosperity", "行业景气度跟踪 HTML 文件不存在，请先到后台上传。")


@app.route("/credit-std-dev")
@login_required
def credit_std_dev():
    local_echarts = url_for("static", filename="vendor/echarts.min.js")
    return html_response(
        STD_DEV_HTML,
        "credit_std_dev",
        "信用债两倍标准差 HTML 文件不存在，请先到后台上传。",
        replacements={
            'src="data/spread_data.js"': f'src="{url_for("credit_std_dev_data_js")}"',
            "src='data/spread_data.js'": f"src='{url_for('credit_std_dev_data_js')}'",
            "https://cdn.jsdelivr.net/npm/echarts@5.5.0/dist/echarts.min.js": local_echarts,
        },
    )


@app.route("/data/credit_std_dev/spread_data.js")
@login_required
def credit_std_dev_data_js():
    if not STD_DEV_JS.exists():
        return Response("var SPREAD_DATA = {categories: [], tenors_by_category: {}, data: {}, update_time: '数据文件未找到'};", mimetype="application/javascript; charset=utf-8")
    return send_file(STD_DEV_JS, mimetype="application/javascript")
@app.route("/admin", methods=["GET", "POST"])
@login_required
@admin_required
def admin():
    message = None
    error = None
    if request.method == "POST":
        target = request.form.get("target", "")
        file = request.files.get("file")
        if not file or not file.filename:
            error = "请选择要上传的文件"
        else:
            try:
                if target == "unified_excel":
                    save_upload(file, juyuan_config.UNIFIED_EXCEL, {".xlsx", ".xls"}, "unified_excel")
                    result = import_unified_excel(juyuan_config.UNIFIED_EXCEL)
                    load_bond_data()
                    message = (
                        "统一 Excel 上传成功："
                        f"债券 {result['bonds']:,} 条，择券候选 {result['bond_picker_bonds']:,} 条，"
                        f"基金指数 {result['fund_prices']:,} 条（{result['fund_start']} 至 {result['fund_end']}）。"
                    )
                elif target == "bond_picker":
                    error = "该上传入口已合并为统一数据 Excel"
                elif target == "strategy_dashboard":
                    error = "策略仪表盘 HTML 上传入口已停用，请使用统一数据 Excel 和一键更新"
                elif target == "spread_monitor_js":
                    error = "利差监控 JS 上传入口已停用，请使用一键更新"
                elif target == "spread_monitor_bond_list":
                    error = "存量债券清单上传入口已合并为统一数据 Excel"
                elif target == "industry_prosperity":
                    save_upload(file, INDUSTRY_HTML, {".html", ".htm"}, "industry_prosperity")
                    message = "行业景气度跟踪 HTML 上传成功。"
                elif target == "credit_std_dev_js":
                    save_upload(file, STD_DEV_JS, {".js"}, "credit_std_dev_js")
                    message = "信用债两倍标准差 spread_data.js 上传成功。"
                else:
                    error = "未知上传目标"
            except Exception as exc:
                error = f"上传失败: {exc}"
    return render_portal_template(
        "admin.html",
        "admin",
        status=status_info(),
        update_status=get_update_status(),
        message=message,
        error=error,
    )


@app.route("/admin/start-db-update", methods=["POST"])
@login_required
@admin_required
def admin_start_db_update():
    modules = request.form.getlist("modules")
    ok, message = start_update(modules or None)
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return {"ok": ok, "message": message, "status": get_update_status()}
    return redirect(url_for("admin"))


@app.route("/admin/update-settings", methods=["POST"])
@login_required
@admin_required
def admin_update_settings():
    try:
        settings = save_update_settings({
            "excel_term_base_date": request.form.get("excel_term_base_date", ""),
        })
        return render_portal_template(
            "admin.html",
            "admin",
            status=status_info(),
            update_status=get_update_status(),
            message=f"Excel 待偿期限基准日已保存为 {settings['excel_term_base_date']}。",
            error=None,
        )
    except Exception as exc:
        return render_portal_template(
            "admin.html",
            "admin",
            status=status_info(),
            update_status=get_update_status(),
            message=None,
            error=f"保存失败: {exc}",
        )



@app.route("/api/status")
@login_required
@admin_required
def api_status():
    return status_info()


@app.route("/api/update-status")
@login_required
@admin_required
def api_update_status():
    return get_update_status()


for directory in [BOND_DIR, STRATEGY_DIR, SPREAD_DIR, INDUSTRY_DIR, STD_DEV_JS.parent, INSTITUTION_FLOW_CACHE.parent, PRIMARY_PRICING_CACHE.parent, CONFIG_DIR, UPLOADS_DIR, juyuan_config.PROJECT2_BOND_EXCEL.parent, juyuan_config.STRATEGY_FUND_PRICES_FROZEN.parent, juyuan_config.BOND_STATIC_JSON.parent, juyuan_config.BOND_PICKER_YIELDS_CACHE.parent]:
    directory.mkdir(parents=True, exist_ok=True)

load_bond_data()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", "5000"))
    debug = os.environ.get("FLASK_DEBUG", "1") == "1"
    app.run(host="0.0.0.0", port=port, debug=debug)
