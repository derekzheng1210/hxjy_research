"""Portal-backed internal-rating lookup for primary-market results."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from juyuan_update import config as portal_config
from juyuan_update.unified_excel import parse_bond_sheet


_CACHE_MTIME_NS: int | None = None
_CACHE: dict[str, str] = {}


def _date_key(value: object) -> str:
    return str(value or "").replace("-", "")[:8]


def _load_ratings(path: Path) -> dict[str, str]:
    if not path.exists():
        return {}
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if "Sheet3" not in workbook.sheetnames:
            return {}
        bonds = parse_bond_sheet(workbook["Sheet3"])
    finally:
        workbook.close()

    latest: dict[str, tuple[str, str]] = {}
    for bond in bonds:
        issuer = str(bond.get("issuer") or "").strip()
        rating = str(bond.get("internal_rating") or "").strip()
        if not issuer or not rating:
            continue
        candidate = (_date_key(bond.get("issue_date")), rating)
        if issuer not in latest or candidate[0] >= latest[issuer][0]:
            latest[issuer] = candidate
    return {issuer: value[1] for issuer, value in latest.items()}


def internal_rating_by_issuer() -> dict[str, str]:
    """Return the latest non-empty portal internal rating for each issuer."""
    global _CACHE_MTIME_NS, _CACHE
    path = portal_config.UNIFIED_EXCEL
    mtime_ns = path.stat().st_mtime_ns if path.exists() else None
    if mtime_ns != _CACHE_MTIME_NS:
        _CACHE = _load_ratings(path)
        _CACHE_MTIME_NS = mtime_ns
    return _CACHE


def attach_internal_ratings(bonds: list[dict]) -> list[dict]:
    ratings = internal_rating_by_issuer()
    for bond in bonds:
        bond["internal_rating"] = ratings.get(str(bond.get("issuer") or "").strip(), "")
    return bonds


def rating_for_issuer(issuer: str) -> str:
    return internal_rating_by_issuer().get(str(issuer or "").strip(), "")
