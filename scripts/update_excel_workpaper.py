"""
一级发行偏离 Excel 底稿更新脚本

功能：
1. 增量刷新 primary_market_pricing/cache.db 到最新日期
2. 基于缓存导出“发行人平均偏离幅度”Excel 底稿

默认行为：
- start_date = 20240101
- end_date = 今天
- 先刷新缓存，再导出 Excel

示例：
    python update_excel_workpaper.py
    python update_excel_workpaper.py --export-only
    python update_excel_workpaper.py --start 20250101 --end 20260715
    python update_excel_workpaper.py --output D:\\tmp\\发行人平均偏离幅度底稿.xlsx
"""

from __future__ import annotations

import argparse
import os
import sqlite3
import sys
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(errors="replace")

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# 项目根加入 sys.path，复用 paths.py 统一数据路径（PORTAL_DATA_ROOT 定位）
_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)
from dotenv import load_dotenv  # noqa: E402
load_dotenv(os.path.join(_PROJECT_ROOT, ".env"))
from paths import WORKPAPER_DIR  # noqa: E402

from primary_market_pricing.cache_builder import (  # noqa: E402
    CACHE_DB_PATH,
    HISTORY_START_DATE,
    ISSUE_DATE_RULE_VERSION,
    DateOrderedCacheBuilder,
    init_cache_db,
)
# 维护脚本已移至 scripts/，请以 `python scripts/update_excel_workpaper.py` 运行。


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
PERCENT_COLUMNS = {"可计算占比", "非市场化占比", "高估占比"}


def reset_cache() -> None:
    """清空当前规则版本缓存。"""
    print("🗑️  清除当前规则版本缓存...")
    conn = init_cache_db(CACHE_DB_PATH)
    try:
        conn.execute(
            "DELETE FROM bond_deviations WHERE issue_date_rule = ?",
            (ISSUE_DATE_RULE_VERSION,),
        )
        conn.execute(
            "DELETE FROM issuer_summary WHERE issue_date_rule = ?",
            (ISSUE_DATE_RULE_VERSION,),
        )
        conn.commit()
    finally:
        conn.close()
    print("   ✅ 已清除")


def refresh_cache(start_date: str, end_date: str, coupon_refresh_days: int) -> None:
    """增量刷新缓存到指定截止日。"""
    builder = DateOrderedCacheBuilder(
        start_date=start_date,
        end_date=end_date,
        coupon_refresh_days=coupon_refresh_days,
    )
    try:
        builder.build()
    finally:
        builder.close()


def _build_summary_df(conn: sqlite3.Connection) -> pd.DataFrame:
    sql = """
    SELECT
        issuer AS 发行人,
        COUNT(*) AS 缓存债券数,
        SUM(CASE WHEN COALESCE(is_no_judgement, 0)=0 AND deviation_bp IS NOT NULL THEN 1 ELSE 0 END) AS 可计算债券数,
        ROUND(
            1.0 * SUM(CASE WHEN COALESCE(is_no_judgement, 0)=0 AND deviation_bp IS NOT NULL THEN 1 ELSE 0 END)
            / COUNT(*),
            4
        ) AS 可计算占比,
        SUM(CASE WHEN COALESCE(is_no_judgement, 0)=0 AND deviation_bp IS NOT NULL AND is_non_market=1 THEN 1 ELSE 0 END) AS 非市场化只数,
        ROUND(
            1.0 * SUM(CASE WHEN COALESCE(is_no_judgement, 0)=0 AND deviation_bp IS NOT NULL AND is_non_market=1 THEN 1 ELSE 0 END)
            / NULLIF(SUM(CASE WHEN COALESCE(is_no_judgement, 0)=0 AND deviation_bp IS NOT NULL THEN 1 ELSE 0 END), 0),
            4
        ) AS 非市场化占比,
        SUM(CASE WHEN COALESCE(is_no_judgement, 0)=0 AND deviation_bp IS NOT NULL AND is_overpriced=1 THEN 1 ELSE 0 END) AS 高估只数,
        ROUND(
            1.0 * SUM(CASE WHEN COALESCE(is_no_judgement, 0)=0 AND deviation_bp IS NOT NULL AND is_overpriced=1 THEN 1 ELSE 0 END)
            / NULLIF(SUM(CASE WHEN COALESCE(is_no_judgement, 0)=0 AND deviation_bp IS NOT NULL THEN 1 ELSE 0 END), 0),
            4
        ) AS 高估占比,
        ROUND(AVG(CASE WHEN COALESCE(is_no_judgement, 0)=0 AND deviation_bp IS NOT NULL THEN deviation_bp END), 2) AS 平均偏离_bp_有符号,
        ROUND(AVG(CASE WHEN COALESCE(is_no_judgement, 0)=0 AND deviation_bp IS NOT NULL THEN ABS(deviation_bp) END), 2) AS 平均偏离幅度_bp,
        ROUND(MIN(CASE WHEN COALESCE(is_no_judgement, 0)=0 AND deviation_bp IS NOT NULL THEN deviation_bp END), 2) AS 最小偏离_bp,
        ROUND(MAX(CASE WHEN COALESCE(is_no_judgement, 0)=0 AND deviation_bp IS NOT NULL THEN deviation_bp END), 2) AS 最大偏离_bp,
        MIN(issue_date) AS 最早发行日,
        MAX(issue_date) AS 最晚发行日,
        ? AS 缓存口径
    FROM bond_deviations
    WHERE issue_date_rule = ?
      AND issuer IS NOT NULL
      AND issuer <> ''
    GROUP BY issuer
    """
    df = pd.read_sql_query(sql, conn, params=(ISSUE_DATE_RULE_VERSION, ISSUE_DATE_RULE_VERSION))
    if df.empty:
        return df

    df["是否有可计算样本"] = df["可计算债券数"].fillna(0).astype(int).gt(0).map({True: "是", False: "否"})
    df = df.sort_values(
        by=["平均偏离幅度_bp", "可计算债券数", "发行人"],
        ascending=[False, False, True],
        na_position="last",
    ).reset_index(drop=True)
    df.insert(0, "偏离幅度排名", df["平均偏离幅度_bp"].rank(method="min", ascending=False, na_option="bottom"))
    df["偏离幅度排名"] = df["偏离幅度排名"].where(df["是否有可计算样本"] == "是")
    df["偏离幅度排名"] = df["偏离幅度排名"].astype("Int64")
    return df


def _build_detail_df(conn: sqlite3.Connection) -> pd.DataFrame:
    sql = """
    SELECT
        symbol AS 债券代码,
        bond_name AS 债券简称,
        issuer AS 发行人,
        issue_date AS 发行日,
        effective_term AS 有效期限_年,
        rating AS 发行评级,
        implied_rating AS 隐含评级,
        effective_rating AS 有效评级,
        raise_mode AS 募集方式,
        bond_type AS 债券类型,
        cvtbd_expire AS 含权期限说明,
        coupon_rate AS 票面利率,
        issue_amount_wan AS 发行规模_万元,
        ref_bond_name AS 参考债简称,
        ref_bond_symbol AS 参考债代码,
        ref_start_date AS 参考估值日期,
        ref_date_gap_years AS 参考日期间隔_年,
        ref_yield AS 参考债估值收益率,
        ref_term AS 参考债估值期限,
        curve_code AS 曲线代码,
        curve_at_ref AS 曲线参考点收益率,
        curve_at_target AS 曲线目标点收益率,
        spread AS 利差_spread,
        fair_price AS 合理发行利率,
        deviation AS 一级偏离,
        deviation_bp AS 偏离_bp,
        ABS(deviation_bp) AS 偏离幅度_bp,
        is_non_market AS 是否非市场化,
        is_overpriced AS 是否高估,
        is_no_judgement AS 是否不可判断,
        computed_at AS 缓存计算时间,
        issue_date_rule AS 缓存口径
    FROM bond_deviations
    WHERE issue_date_rule = ?
    ORDER BY issuer, issue_date, symbol
    """
    return pd.read_sql_query(sql, conn, params=(ISSUE_DATE_RULE_VERSION,))


def _build_meta_df(
    db_path: str,
    output_path: str,
    summary_df: pd.DataFrame,
    detail_df: pd.DataFrame,
    refreshed: bool,
    start_date: str,
    end_date: str,
) -> pd.DataFrame:
    issuer_cnt = int(len(summary_df))
    issuer_with_calc = int((summary_df.get("是否有可计算样本") == "是").sum()) if not summary_df.empty else 0
    cache_bonds_total = int(summary_df["缓存债券数"].fillna(0).sum()) if not summary_df.empty else 0
    calc_bonds_total = int(summary_df["可计算债券数"].fillna(0).sum()) if not summary_df.empty else 0
    min_issue_date = detail_df["发行日"].min() if not detail_df.empty else ""
    max_issue_date = detail_df["发行日"].max() if not detail_df.empty else ""

    rows = [
        ["缓存文件", db_path],
        ["导出文件", output_path],
        ["提取时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["是否先刷新缓存", "是" if refreshed else "否（仅导出已有缓存）"],
        ["本次刷新起始日期", start_date],
        ["本次刷新截止日期", end_date],
        ["使用缓存口径", ISSUE_DATE_RULE_VERSION],
        ["缓存覆盖发行日区间", f"{min_issue_date} ~ {max_issue_date}" if min_issue_date and max_issue_date else ""],
        ["发行人数量", issuer_cnt],
        ["其中有可计算样本的发行人数量", issuer_with_calc],
        ["缓存债券总数", cache_bonds_total],
        ["可计算债券总数", calc_bonds_total],
        ["平均偏离幅度_bp 定义", "按发行人对可计算债券的 ABS(deviation_bp) 求平均"],
        ["平均偏离_bp_有符号 定义", "按发行人对可计算债券的 deviation_bp 求平均，保留方向"],
        ["非市场化判定阈值", "低于合理价格 3BP 以上判定为非市场化；高于合理价格 3BP 以上判定为高估"],
        ["说明", "为避免“偏离幅度”歧义，汇总表同时保留平均偏离幅度_bp（绝对值）和平均偏离_bp_有符号两列"],
    ]
    return pd.DataFrame(rows, columns=["字段", "值"])


def _style_workbook(output_path: str) -> None:
    wb = load_workbook(output_path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT

        header_map = {cell.value: cell.column for cell in ws[1]}
        for col_name in PERCENT_COLUMNS:
            col_idx = header_map.get(col_name)
            if not col_idx:
                continue
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "0.00%"

        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 28)

    wb.save(output_path)


def export_workbook(output_path: str, refreshed: bool, start_date: str, end_date: str) -> str:
    """从缓存导出 Excel 底稿。"""
    if not os.path.exists(CACHE_DB_PATH):
        raise FileNotFoundError(f"未找到缓存文件：{CACHE_DB_PATH}")

    conn = sqlite3.connect(CACHE_DB_PATH)
    try:
        summary_df = _build_summary_df(conn)
        detail_df = _build_detail_df(conn)
        meta_df = _build_meta_df(
            db_path=CACHE_DB_PATH,
            output_path=output_path,
            summary_df=summary_df,
            detail_df=detail_df,
            refreshed=refreshed,
            start_date=start_date,
            end_date=end_date,
        )
    finally:
        conn.close()

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="发行人汇总", index=False)
        detail_df.to_excel(writer, sheet_name="债券明细", index=False)
        meta_df.to_excel(writer, sheet_name="口径说明", index=False)

    _style_workbook(output_path)
    return output_path


def build_default_output_path() -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_dir = str(WORKPAPER_DIR)
    os.makedirs(base_dir, exist_ok=True)
    return os.path.join(base_dir, f"发行人平均偏离幅度底稿_{ts}.xlsx")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="刷新 primary_market_pricing 缓存并导出 Excel 底稿")
    parser.add_argument(
        "--start",
        default=HISTORY_START_DATE,
        help=f"起始日期，格式 YYYYMMDD（默认：{HISTORY_START_DATE}）",
    )
    parser.add_argument(
        "--end",
        default=None,
        help="截止日期，格式 YYYYMMDD（默认：今天）",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="输出 Excel 路径（默认：项目目录下自动生成时间戳文件名）",
    )
    parser.add_argument(
        "--export-only",
        action="store_true",
        help="只导出 Excel，不刷新缓存",
    )
    parser.add_argument(
        "--reset-cache",
        action="store_true",
        help="导出前清空当前规则版本缓存后重算",
    )
    parser.add_argument(
        "--coupon-refresh-days",
        type=int,
        default=7,
        help="近期票面利率缺失的回补窗口天数（默认：7）",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    end_date = args.end or datetime.now().strftime("%Y%m%d")
    output_path = args.output or build_default_output_path()

    if args.export_only and args.reset_cache:
        raise ValueError("--export-only 与 --reset-cache 不能同时使用")

    refreshed = False
    if args.reset_cache:
        reset_cache()

    if not args.export_only:
        print(f"🚀 开始刷新缓存：{args.start} ~ {end_date}")
        refresh_cache(args.start, end_date, args.coupon_refresh_days)
        refreshed = True
    else:
        print("📦 跳过缓存刷新，直接导出已有缓存")

    print("📄 开始导出 Excel 底稿...")
    final_path = export_workbook(output_path, refreshed, args.start, end_date)
    print(f"✅ 底稿已生成：{final_path}")


if __name__ == "__main__":
    main()
