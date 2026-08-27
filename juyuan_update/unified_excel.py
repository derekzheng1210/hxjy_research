from __future__ import annotations

import json
import os
import re
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

from . import config


RATING_RANK = {
    "AAA+": 12,
    "AAA": 11,
    "AAA-": 10,
    "AA+": 9,
    "AA": 8,
    "AA(2)": 8,
    "AA-": 7,
    "A+": 6,
    "A": 5,
    "A-": 4,
    "BBB+": 3,
    "BBB": 2,
    "BBB-": 1,
}

DEFAULT_EXCEL_TERM_BASE_DATE = "2026-06-10"


def atomic_write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(text, encoding="utf-8", newline="\n")
    os.replace(tmp, path)


def normalize_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() in {"nan", "none"} else text


def normalize_bond_code(code) -> str:
    text = normalize_text(code).replace(" ", "").upper()
    if not text:
        return ""
    if text.endswith((".IB", ".SH", ".SZ", ".BJ")):
        return text
    if not text.isdigit():
        return text
    if len(text) >= 9:
        return text + ".IB"
    if len(text) == 6:
        if text[:2] in ("52",) or text[:3] in ("148", "149", "111", "112"):
            return text + ".SZ"
        return text + ".SH"
    return text + ".IB"


def normalize_rating(value) -> str:
    text = normalize_text(value).upper().replace(" ", "")
    if text in {"#N/A", "NA", "N/A", "-", "--"}:
        return ""
    return text


def rating_at_least(value, floor: str = "BBB-") -> bool:
    rating = normalize_rating(value)
    return RATING_RANK.get(rating, 0) >= RATING_RANK[floor]


def is_blank(value) -> bool:
    return normalize_text(value) == ""


def is_yes(value) -> bool:
    return normalize_text(value).lower() in {"是", "yes", "y", "1", "true"}


def date_text(value) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)) and 1 <= value <= 60000:
        # Excel serial dates use 1899-12-30 as the practical epoch in openpyxl.
        return (datetime(1899, 12, 30) + timedelta(days=int(value))).strftime("%Y-%m-%d")
    text = normalize_text(value)
    if not text:
        return ""
    text = (
        text.replace("年", "-")
        .replace("月", "-")
        .replace("日", "")
        .replace(".", "-")
        .replace("/", "-")
    )
    text = re.sub(r"\s*-\s*", "-", text)
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(text[:19], fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    if re.match(r"^\d{8}$", text):
        return f"{text[:4]}-{text[4:6]}-{text[6:8]}"
    match = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})", text)
    if match:
        year, month, day = match.groups()
        return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"
    return text[:10]


_json_cache: dict[Path, tuple[float, float, object]] = {}


def _load_json_cached(path: Path, default):
    """Read and parse *path* as JSON, memoised by (mtime, size).

    Repeated calls within the same process skip disk I/O and JSON parsing
    when the file has not changed. Writes via ``write_json`` / ``atomic_write_text``
    update the file's mtime, so the cache invalidates automatically.
    """
    try:
        stat = path.stat()
    except OSError:
        _json_cache.pop(path, None)
        return default
    cached = _json_cache.get(path)
    if cached and cached[0] == stat.st_mtime and cached[1] == stat.st_size:
        return cached[2]
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        _json_cache.pop(path, None)
        return default
    _json_cache[path] = (stat.st_mtime, stat.st_size, payload)
    return payload


def load_json(path: Path, default):
    return _load_json_cached(path, default)


def write_json(path: Path, payload) -> None:
    atomic_write_text(path, json.dumps(payload, ensure_ascii=False, indent=2))
    _json_cache.pop(path, None)


def normalize_date(value, *, field_name: str = "日期") -> str:
    text = date_text(value)
    try:
        return datetime.strptime(text, "%Y-%m-%d").strftime("%Y-%m-%d")
    except ValueError as exc:
        raise ValueError(f"{field_name}必须是 YYYY-MM-DD 格式") from exc


def load_update_settings() -> dict:
    payload = load_json(config.UPDATE_SETTINGS_JSON, {})
    base_date = payload.get("excel_term_base_date") or DEFAULT_EXCEL_TERM_BASE_DATE
    try:
        base_date = normalize_date(base_date, field_name="Excel 待偿期限基准日")
    except ValueError:
        base_date = DEFAULT_EXCEL_TERM_BASE_DATE
    return {
        "excel_term_base_date": base_date,
    }


def save_update_settings(settings: dict) -> dict:
    payload = load_update_settings()
    if "excel_term_base_date" in settings:
        payload["excel_term_base_date"] = normalize_date(
            settings["excel_term_base_date"],
            field_name="Excel 待偿期限基准日",
        )
    write_json(config.UPDATE_SETTINGS_JSON, payload)
    return payload


def load_bond_static() -> dict:
    return load_json(config.BOND_STATIC_JSON, {"generated_at": "", "source_file": "", "bonds": []})


def save_bond_static(bonds: list[dict], source_file: str) -> dict:
    payload = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "source_file": source_file,
        "total_bonds": len(bonds),
        "bonds": bonds,
    }
    write_json(config.BOND_STATIC_JSON, payload)
    return payload


def load_counterparty_limits() -> dict:
    return load_json(
        config.COUNTERPARTY_LIMITS_JSON,
        {"generated_at": "", "source_file": "", "total_issuers": 0, "limits": {}},
    )


def save_counterparty_limits(limits: dict[str, float], source_file: str) -> dict:
    payload = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "source_file": source_file,
        "total_issuers": len(limits),
        "limits": limits,
    }
    write_json(config.COUNTERPARTY_LIMITS_JSON, payload)
    return payload


def refresh_bond_terms(target_trade_date: str, base_date: str | None = None) -> dict:
    import copy
    payload = copy.deepcopy(load_bond_static())
    bonds = payload.get("bonds") or []
    if not bonds:
        return {
            "base_date": base_date or load_update_settings()["excel_term_base_date"],
            "target_trade_date": normalize_date(target_trade_date, field_name="目标交易日"),
            "updated": 0,
            "total_bonds": 0,
        }

    base_text = normalize_date(base_date or load_update_settings()["excel_term_base_date"], field_name="Excel 待偿期限基准日")
    target_text = normalize_date(target_trade_date, field_name="目标交易日")
    base_dt = datetime.strptime(base_text, "%Y-%m-%d").date()
    target_dt = datetime.strptime(target_text, "%Y-%m-%d").date()

    updated = 0
    for bond in bonds:
        try:
            original_term = float(bond.get("original_term", bond.get("term")))
        except (TypeError, ValueError):
            continue
        if original_term < 0:
            continue
        maturity_dt = base_dt + timedelta(days=round(original_term * 365))
        new_term = max((maturity_dt - target_dt).days / 365, 0)
        bond["original_term"] = round(original_term, 4)
        bond["term"] = round(new_term, 4)
        bond["term_base_date"] = base_text
        bond["term_updated_date"] = target_text
        bond["maturity_date_estimated"] = maturity_dt.strftime("%Y-%m-%d")
        updated += 1

    payload["generated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    payload["term_base_date"] = base_text
    payload["term_updated_date"] = target_text
    payload["total_bonds"] = len(bonds)
    payload["bonds"] = bonds
    write_json(config.BOND_STATIC_JSON, payload)
    return {
        "base_date": base_text,
        "target_trade_date": target_text,
        "updated": updated,
        "total_bonds": len(bonds),
    }


def get_bond_picker_bonds() -> list[dict]:
    return [
        bond for bond in load_bond_static().get("bonds", [])
        if rating_at_least(bond.get("internal_rating")) and is_blank(bond.get("guarantor"))
    ]


def get_spread_monitor_bonds() -> list[dict]:
    return list(load_bond_static().get("bonds", []))


def load_bond_picker_yields_cache() -> dict:
    return load_json(config.BOND_PICKER_YIELDS_CACHE, {"trade_date": "", "generated_at": "", "yields": {}})


def save_bond_picker_yields_cache(trade_date: str, yields: dict[str, float]) -> dict:
    normalized_yields = {}
    for key, value in yields.items():
        if value is None:
            continue
        raw = str(key or "").strip().upper()
        if not raw:
            continue
        normalized_yields[raw] = float(value)
        bare = raw.split(".", 1)[0]
        normalized_yields.setdefault(bare, float(value))
    payload = {
        "trade_date": trade_date,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "yields": normalized_yields,
    }
    write_json(config.BOND_PICKER_YIELDS_CACHE, payload)
    return payload


def load_spread_history_cache() -> dict:
    return load_json(config.SPREAD_HISTORY_CACHE, {"generated_at": "", "dates": {}})


def save_spread_history_cache(cache: dict) -> None:
    cache["generated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    write_json(config.SPREAD_HISTORY_CACHE, cache)


def _header_index(headers: list[str], candidates: list[str], default=None):
    for candidate in candidates:
        for idx, header in enumerate(headers):
            if candidate == header or candidate in header:
                return idx
    return default


def parse_bond_sheet(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [normalize_text(h) for h in rows[0]]

    idx = {
        "code": _header_index(headers, ["证券代码", "债券代码", "代码"], 0),
        "name": _header_index(headers, ["证券简称", "证券名称", "债券名称", "名称"], 1),
        "term": _header_index(headers, ["待偿年限", "待偿期限", "剩余期限", "期限"], 2),
        "issue_date": _header_index(headers, ["起息日期", "起息日"], None),
        "implied_rating": _header_index(headers, ["隐含评级"], None),
        "issuer": _header_index(headers, ["债务主体中文名称", "融资主体", "主体名称"], None),
        "entity": _header_index(headers, ["主体性质"], None),
        "ct": _header_index(headers, ["是否城投债", "城投债"], None),
        "sub": _header_index(headers, ["是否次级债", "次级债"], None),
        "tech": _header_index(headers, ["是否科创债", "科创债"], None),
        "guarantor": _header_index(headers, ["担保人"], None),
        "internal_rating": _header_index(headers, ["内评", "内部评级"], None),
        "holding": _header_index(headers, ["是否持仓", "持仓"], None),
    }
    required = ["code", "name", "term"]
    missing = [name for name in required if idx[name] is None]
    if missing:
        raise RuntimeError("统一 Excel 的 Sheet3 缺少关键列: " + ", ".join(missing))

    bonds = []
    seen = set()
    for row in rows[1:]:
        code = normalize_bond_code(row[idx["code"]] if idx["code"] < len(row) else None)
        if not code or code in seen:
            continue
        try:
            term = float(row[idx["term"]])
        except Exception:
            continue
        if term <= 0:
            continue
        seen.add(code)

        def value(name):
            col = idx.get(name)
            return normalize_text(row[col]) if col is not None and col < len(row) else ""

        bonds.append({
            "code": code,
            "raw_code": normalize_text(row[idx["code"]] if idx["code"] < len(row) else code),
            "name": value("name"),
            "term": round(term, 4),
            "issue_date": date_text(row[idx["issue_date"]]) if idx["issue_date"] is not None and idx["issue_date"] < len(row) else "",
            "implied_rating": normalize_rating(value("implied_rating")),
            "issuer": value("issuer"),
            "entity": value("entity"),
            "ct": value("ct"),
            "sub": value("sub"),
            "tech": value("tech"),
            "guarantor": value("guarantor"),
            "internal_rating": normalize_rating(value("internal_rating")),
            "is_holding": is_yes(value("holding")),
        })
    return bonds


def parse_fund_sheet(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    fund = []
    for row in rows:
        if len(row) < 3:
            continue
        dt = date_text(row[1])
        try:
            close = float(row[2])
        except Exception:
            continue
        if re.match(r"^\d{4}-\d{2}-\d{2}$", dt) and close > 0:
            fund.append({"date": dt, "close": round(close, 6)})
    fund.sort(key=lambda x: x["date"])
    dedup = {}
    for row in fund:
        dedup[row["date"]] = row
    return [dedup[d] for d in sorted(dedup)]


def parse_neiping_sheet(ws) -> dict[str, float]:
    rows = list(ws.iter_rows(values_only=True))
    header_row = None
    issuer_index = None
    limit_index = None
    for row_index, row in enumerate(rows[:20]):
        headers = [normalize_text(value) for value in row]
        issuer_index = _header_index(headers, ["融资主体", "主体名称"], None)
        limit_index = _header_index(headers, ["最新可用对手限额"], None)
        if issuer_index is not None and limit_index is not None:
            header_row = row_index
            break
    if header_row is None:
        raise RuntimeError("统一 Excel 的 neiping 缺少融资主体或最新可用对手限额列")

    limits: dict[str, float] = {}
    for row in rows[header_row + 1:]:
        issuer = normalize_text(row[issuer_index] if issuer_index < len(row) else None)
        if not issuer:
            continue
        try:
            value = float(row[limit_index])
        except (TypeError, ValueError, IndexError):
            continue
        if value != value or value in (float("inf"), float("-inf")):
            continue
        limits[issuer] = round(value, 6)
    return limits


def import_unified_excel(path: Path) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if "Sheet3" not in wb.sheetnames:
            raise RuntimeError("统一 Excel 缺少 Sheet3 债券信息表")
        if "Sheet1" not in wb.sheetnames:
            raise RuntimeError("统一 Excel 缺少 Sheet1 基金指数表")
        if "neiping" not in wb.sheetnames:
            raise RuntimeError("统一 Excel 缺少 neiping 内评与限额表")
        bonds = parse_bond_sheet(wb["Sheet3"])
        fund_prices = parse_fund_sheet(wb["Sheet1"])
        counterparty_limits = parse_neiping_sheet(wb["neiping"])
    finally:
        wb.close()

    if not bonds:
        raise RuntimeError("Sheet3 未读取到有效债券数据")
    if not fund_prices:
        raise RuntimeError("Sheet1 未读取到有效基金指数数据")
    if not counterparty_limits:
        raise RuntimeError("neiping 未读取到有效对手限额数据")

    bond_payload = save_bond_static(bonds, str(path))
    limit_payload = save_counterparty_limits(counterparty_limits, str(path))
    write_json(config.STRATEGY_FUND_PRICES_FROZEN, fund_prices)
    return {
        "bonds": bond_payload["total_bonds"],
        "bond_picker_bonds": len(get_bond_picker_bonds()),
        "fund_prices": len(fund_prices),
        "fund_start": fund_prices[0]["date"],
        "fund_end": fund_prices[-1]["date"],
        "counterparty_limit_issuers": limit_payload["total_issuers"],
    }
