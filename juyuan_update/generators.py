from __future__ import annotations

import json
import os
import re
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

from . import config
from .db import (
    connect,
    fetch_cnbd_yields_by_symbol,
    fetch_curve_series,
    fetch_curve_series_for_dates,
    latest_curve_date,
    resolve_curve_codes,
    shclest_reference_dates,
)
from .strategy_dashboard import build_dashboard as build_strategy_dashboard
from .unified_excel import (
    get_bond_picker_bonds,
    get_spread_monitor_bonds,
    load_update_settings,
    load_spread_history_cache,
    refresh_bond_terms,
    save_bond_picker_yields_cache,
    save_spread_history_cache,
)


TENORS = list(config.STD_DEV_TENORS)


def tenor_key(tenor: int | float) -> str:
    value = float(tenor)
    if value.is_integer():
        return str(int(value))
    return f"{value:g}"


def tenor_label(tenor: int | float) -> str:
    if float(tenor) == float(config.ONE_MONTH_TENOR):
        return "1M"
    return f"{tenor_key(tenor)}Y"


def atomic_write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(text, encoding="utf-8", newline="\n")
    os.replace(tmp, path)


def write_js(path: Path, data: dict) -> None:
    atomic_write_text(path, "var SPREAD_DATA = " + json.dumps(data, ensure_ascii=False) + ";\n")


def read_js_data(path: Path) -> dict | None:
    if not path.exists():
        return None
    text = path.read_text(encoding="utf-8", errors="replace").strip()
    match = re.search(r"var\s+SPREAD_DATA\s*=\s*(\{.*\})\s*;?\s*$", text, flags=re.S)
    if not match:
        return None
    return json.loads(match.group(1))


def dash_date(value: str) -> str:
    text = str(value).replace("-", "")[:8]
    return f"{text[:4]}-{text[4:6]}-{text[6:]}"


def _date_obj(value: str):
    return datetime.strptime(str(value).replace("-", "")[:8], "%Y%m%d").date()


def _nearest_cached_date(cached_dates: dict, target_date: str, before: bool = False) -> str | None:
    if not cached_dates:
        return None
    target = str(target_date).replace("-", "")[:8]
    candidates = []
    for dt, snapshot in cached_dates.items():
        dt_text = str(dt)
        if not snapshot:
            continue
        if (before and dt_text < target) or (not before and dt_text <= target):
            candidates.append(dt_text)
    return max(candidates) if candidates else None


def cached_reference_dates(cached_dates: dict, target_date: str) -> dict[str, str]:
    current = _nearest_cached_date(cached_dates, target_date)
    if not current:
        return {}
    current_dt = _date_obj(current)
    return {
        label: dt
        for label, dt in {
            "当前": current,
            "昨日": _nearest_cached_date(cached_dates, current, before=True),
            "一周前": _nearest_cached_date(cached_dates, (current_dt - timedelta(days=7)).strftime("%Y%m%d")),
            "一月前": _nearest_cached_date(cached_dates, (current_dt - timedelta(days=30)).strftime("%Y%m%d")),
            "年初": _nearest_cached_date(cached_dates, f"{current_dt.year}0101", before=True),
        }.items()
        if dt
    }


def normalize_reference_date_labels(dates: dict) -> dict[str, str]:
    labels = ["当前", "昨日", "一周前", "一月前", "年初"]
    normalized = {label: dates[label] for label in labels if label in dates}
    if len(normalized) == len(dates):
        return normalized
    for label, key in zip(labels, dates.keys()):
        normalized.setdefault(label, dates[key])
    return normalized


def load_std_dev_latest_curve_cache() -> dict:
    if not config.STD_DEV_CURVES_CACHE.exists():
        return {}
    try:
        return json.loads(config.STD_DEV_CURVES_CACHE.read_text(encoding="utf-8")).get("curves_by_date", {})
    except Exception:
        return {}


def save_std_dev_latest_curve_cache(curves_by_date: dict, oracle_latest: str) -> None:
    existing = load_std_dev_latest_curve_cache()
    existing.update(curves_by_date)
    dates = sorted(existing)
    payload = {
        "start_date": dates[0] if dates else dash_date(oracle_latest),
        "end_date": dates[-1] if dates else dash_date(oracle_latest),
        "oracle_latest": dash_date(oracle_latest),
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "curves_by_date": existing,
    }
    atomic_write_text(config.STD_DEV_CURVES_CACHE, json.dumps(payload, ensure_ascii=False))


def fetch_std_dev_latest_curves(conn, latest_date: str, curve_codes: dict[str, dict]) -> dict:
    latest_dash = dash_date(latest_date)
    cached = load_std_dev_latest_curve_cache()
    required_tenors = {tenor_key(tenor) for tenor in TENORS}
    cached_day = cached.get(latest_dash)
    if cached_day and all(
        required_tenors.issubset(set((cached_day.get(curve_key) or {}).keys()))
        for curve_key in curve_codes
    ):
        return cached_day
    series = fetch_curve_series_for_dates(conn, curve_codes, [latest_date], tenors=TENORS)
    day: dict[str, dict[str, float]] = {}
    for curve_key, by_tenor in series.items():
        values = {}
        for tenor in TENORS:
            raw = by_tenor.get(float(tenor), {}).get(latest_date)
            if raw is not None:
                values[tenor_key(tenor)] = float(raw)
        day[curve_key] = values
    if not any(day.values()):
        raise RuntimeError(f"Oracle 最新曲线结果为空: {latest_dash}")
    save_std_dev_latest_curve_cache({latest_dash: day}, latest_date)
    return day


def merge_spread_point(payload: dict, key: str, date_str: str, value: float) -> None:
    entry = payload["data"].setdefault(key, {"dates": [], "values": []})
    dates = [str(d)[:10] for d in entry.get("dates", [])]
    values = list(entry.get("values", []))
    rounded = round(value, 6)
    if date_str in dates:
        idx = dates.index(date_str)
        values[idx] = rounded
    else:
        dates.append(date_str)
        values.append(rounded)
    ordered = sorted(zip(dates, values), key=lambda item: item[0])
    entry["dates"] = [d for d, _ in ordered]
    entry["values"] = [v for _, v in ordered]


def std_dev_history_start(data: dict) -> str | None:
    for key, series in data.get("data", {}).items():
        if str(key).endswith("_1Y"):
            dates = series.get("dates") or []
            if dates:
                return str(dates[0]).replace("-", "")[:8]
    return None


def std_dev_recent_dates(data: dict, limit: int = 90) -> list[str]:
    for key, series in data.get("data", {}).items():
        if str(key).endswith("_1Y"):
            dates = [str(dt).replace("-", "")[:8] for dt in series.get("dates") or []]
            return dates[-limit:]
    return []


def std_dev_reference_dates(data: dict) -> list[str]:
    for key, series in data.get("data", {}).items():
        if str(key).endswith("_1Y"):
            return [str(dt).replace("-", "")[:8] for dt in series.get("dates") or []]
    return []


def backfill_std_dev_tenor_history(
    conn,
    data: dict,
    curves: dict[str, dict],
    end_date: str,
    tenor: int | float,
) -> None:
    label_tenor = tenor_label(tenor)
    reference_dates = std_dev_reference_dates(data)
    end_ymd = str(end_date).replace("-", "")[:8]
    if end_ymd not in reference_dates:
        reference_dates.append(end_ymd)
    if not reference_dates:
        reference_dates = [end_ymd]

    target_dates = set(reference_dates)
    existing_dates = set()
    for label, _, _ in config.STD_DEV_SPREADS:
        series_dates = (data["data"].get(f"{label}_{label_tenor}") or {}).get("dates") or []
        existing_dates.update(str(dt).replace("-", "")[:8] for dt in series_dates)
    missing_dates = sorted(target_dates - existing_dates)
    if not missing_dates:
        return

    series = {key: {} for key in curves}
    for start in range(0, len(missing_dates), 20):
        chunk_dates = missing_dates[start:start + 20]
        chunk = fetch_curve_series_for_dates(conn, curves, chunk_dates, tenors=[tenor])
        for curve_key, by_tenor in chunk.items():
            target = series.setdefault(curve_key, {})
            for tenor_value, values in by_tenor.items():
                target.setdefault(tenor_value, {}).update(values)
    key_tenor = tenor_key(tenor)
    for label, credit_key, base_key in config.STD_DEV_SPREADS:
        credit_values = series.get(credit_key, {}).get(float(tenor), {})
        base_values = series.get(base_key, {}).get(float(tenor), {})
        for dt in sorted(set(credit_values) & set(base_values)):
            credit_val = credit_values.get(dt)
            base_val = base_values.get(dt)
            if credit_val is None or base_val is None:
                continue
            merge_spread_point(data, f"{label}_{label_tenor}", dash_date(dt), credit_val - base_val)


def normalize_bond_code(code) -> str:
    text = str(code or "").strip().replace(" ", "")
    if not text:
        return ""
    upper = text.upper()
    if upper.endswith((".IB", ".SH", ".SZ", ".BJ")):
        return upper
    if not text.isdigit():
        return text
    if len(text) >= 9:
        return text + ".IB"
    if len(text) == 6:
        if text[:2] in ("52",) or text[:3] in ("148", "149", "111", "112"):
            return text + ".SZ"
        return text + ".SH"
    return text + ".IB"


def load_project2_bond_list(path: Path | None = None) -> list[dict]:
    if path is None:
        unified_bonds = get_spread_monitor_bonds()
        if unified_bonds:
            return [
                {
                    "code": bond["code"],
                    "raw_code": bond.get("raw_code") or bond["code"],
                    "name": bond.get("name") or "",
                    "term": bond["term"],
                    "implied_rating": bond.get("implied_rating") or "",
                    "is_holding": bool(bond.get("is_holding")),
                }
                for bond in unified_bonds
            ]
    path = path or config.PROJECT2_BOND_EXCEL
    if not path.exists():
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [str(h).strip() if h else "" for h in rows[0]]
    col = {h: i for i, h in enumerate(headers) if h}

    def pick(names, default=None):
        for name in names:
            if name in col:
                return col[name]
        return default

    code_col = pick(["债券代码", "证券代码", "代码"], 0)
    name_col = pick(["证券名称", "债券名称", "名称"], 1)
    term_col = pick(["待偿期限", "剩余期限", "期限"], 2)
    rating_col = pick(["隐含评级", "评级"], None)
    hold_col = pick(["是否持仓", "持仓标记", "持仓"], None)

    bonds = []
    for row in rows[1:]:
        raw_code = row[code_col] if code_col is not None and code_col < len(row) else None
        code = normalize_bond_code(raw_code)
        if not code:
            continue
        try:
            term = float(row[term_col])
        except Exception:
            continue
        if term <= 0:
            continue
        name = str(row[name_col]).strip() if name_col is not None and name_col < len(row) and row[name_col] else ""
        rating = str(row[rating_col]).strip() if rating_col is not None and rating_col < len(row) and row[rating_col] else ""
        is_holding = False
        if hold_col is not None and hold_col < len(row) and row[hold_col] is not None:
            is_holding = str(row[hold_col]).strip().lower() in ("是", "yes", "y", "1", "true")
        bonds.append({
            "code": code,
            "raw_code": str(raw_code).strip(),
            "name": name,
            "term": round(term, 4),
            "implied_rating": rating,
            "is_holding": is_holding,
        })
    return bonds


def interpolate_curve(curve_by_tenor: dict[float, dict[str, float]], tenor: float, date_str: str) -> float | None:
    points = sorted((t, vals.get(date_str)) for t, vals in curve_by_tenor.items() if vals.get(date_str) is not None)
    points = [(t, v) for t, v in points if v is not None]
    if not points:
        return None
    if tenor <= points[0][0]:
        return points[0][1]
    if tenor >= points[-1][0]:
        return points[-1][1]
    for (t1, y1), (t2, y2) in zip(points, points[1:]):
        if t1 <= tenor <= t2:
            if t1 == t2:
                return y1
            ratio = (tenor - t1) / (t2 - t1)
            return round(y1 + ratio * (y2 - y1), 6)
    return points[-1][1]


def interpolate_curve_with_long_end(
    curve_by_tenor: dict[float, dict[str, float]],
    long_curve_by_tenor: dict[float, dict[str, float]],
    tenor: float,
    date_str: str,
) -> float | None:
    points = sorted((t, vals.get(date_str)) for t, vals in curve_by_tenor.items() if vals.get(date_str) is not None)
    points = [(t, v) for t, v in points if v is not None]
    if not points:
        return None

    max_tenor = points[-1][0]
    if tenor <= max_tenor:
        return interpolate_curve(curve_by_tenor, tenor, date_str)

    base_yield = points[-1][1]
    if not long_curve_by_tenor:
        return base_yield

    long_base_yield = interpolate_curve(long_curve_by_tenor, max_tenor, date_str)
    long_target_yield = interpolate_curve(long_curve_by_tenor, tenor, date_str)
    if long_base_yield is None or long_target_yield is None:
        return base_yield

    return round(base_yield + (long_target_yield - long_base_yield), 6)


def curve_for_bond(bond: dict) -> str:
    name = bond.get("name") or ""
    if "二级" in name or "资本" in name:
        return "股份行二级资本债"
    return config.RATING_TO_CURVE.get((bond.get("implied_rating") or "").strip(), "中短票AAA")


def bond_yield_for(yields: dict, code: str):
    raw = str(code or "").strip().upper()
    bare = raw.split(".", 1)[0]
    return yields.get(raw, yields.get(bare))


def generate_std_dev_data(progress=None) -> dict:
    if progress:
        progress("生成两倍标准差最新曲线利差", 5)
    existing = read_js_data(config.STD_DEV_JS) or {
        "categories": [],
        "tenors_by_category": {},
        "data": {},
    }
    data = {
        "categories": list(existing.get("categories") or []),
        "tenors_by_category": dict(existing.get("tenors_by_category") or {}),
        "data": {
            key: value
            for key, value in dict(existing.get("data") or {}).items()
            if not str(key).endswith("_0Y")
        },
        "update_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    with connect() as conn:
        end_date = latest_curve_date(conn)
        latest_dash = dash_date(end_date)
        curves = resolve_curve_codes(conn, config.CURVE_DEFS, end_date)
        missing = [name for name in config.CURVE_DEFS if name not in curves]
        if missing:
            raise RuntimeError("未找到曲线: " + ", ".join(missing))
        if progress:
            progress(f"读取 Oracle 两倍标准差最新曲线 {latest_dash}", 8)
        latest_curves = fetch_std_dev_latest_curves(conn, end_date, curves)
        if progress:
            progress("补齐两倍标准差 1M 历史曲线", 20)
        backfill_std_dev_tenor_history(conn, data, curves, end_date, config.ONE_MONTH_TENOR)

    for label, credit_key, base_key in config.STD_DEV_SPREADS:
        if label not in data["categories"]:
            data["categories"].append(label)
        tenors = []
        for tenor in TENORS:
            key_tenor = tenor_key(tenor)
            credit_val = latest_curves.get(credit_key, {}).get(key_tenor)
            base_val = latest_curves.get(base_key, {}).get(key_tenor)
            if credit_val is None or base_val is None:
                continue
            label_tenor = tenor_label(tenor)
            key = f"{label}_{label_tenor}"
            merge_spread_point(data, key, latest_dash, credit_val - base_val)
            if label_tenor not in tenors:
                tenors.append(label_tenor)
        data["tenors_by_category"][label] = tenors
    write_js(config.STD_DEV_JS, data)
    return data

def generate_spread_monitor(progress=None) -> dict:
    if progress:
        progress("读取存量债券清单", 12)
    bonds = load_project2_bond_list()
    if not bonds:
        raise RuntimeError(f"未找到或无法读取存量债券清单: {config.PROJECT2_BOND_EXCEL}")

    with connect() as conn:
        if progress:
            progress("读取曲线与参考日期", 18)
        end_date = latest_curve_date(conn)
        cache = load_spread_history_cache()
        cached_dates = cache.setdefault("dates", {})
        dates = normalize_reference_date_labels(shclest_reference_dates(conn, end_date))
        if not dates:
            dates = cached_reference_dates(cached_dates, end_date)
        if not dates:
            raise RuntimeError("未查询到中债估值交易日，且历史缓存为空")
        curves = resolve_curve_codes(conn, config.CURVE_DEFS, dates["当前"])
        missing_dates = [dt for dt in dict.fromkeys(dates.values()) if dt and dt not in cached_dates]
        if progress:
            hit_count = len([dt for dt in dates.values() if dt in cached_dates])
            progress(f"读取利差历史缓存，命中 {hit_count} 个日期", 24)

        query_dates = list(dict.fromkeys(dt for dt in dates.values() if dt))
        curve_series = fetch_curve_series_for_dates(conn, curves, query_dates, tenors=config.SPREAD_MONITOR_TENORS)

        yield_data = {}
        labels = list(dates.items())
        for idx, (label, dt) in enumerate(labels, start=1):
            if dt in cached_dates and label != "当前":
                if progress:
                    progress(f"读取利差历史缓存 {label} {dt}", 35 + int(idx / len(labels) * 35))
                continue
            if progress:
                progress(f"查询中债估值收益率 {label} {dt}", 35 + int(idx / len(labels) * 35))
            yield_data[label] = fetch_cnbd_yields_by_symbol(conn, [b["code"] for b in bonds], dt)
            if label == "当前" and not yield_data[label]:
                fallback_dates = cached_reference_dates(cached_dates, dt)
                fallback_current = fallback_dates.get("当前")
                if fallback_current and fallback_current != dt:
                    if progress:
                        progress(f"当前估值 {dt} 为空，回退至历史缓存 {fallback_current}", 60)
                    dates = fallback_dates
                    curve_series = fetch_curve_series_for_dates(
                        conn,
                        curves,
                        list(dict.fromkeys(d for d in dates.values() if d)),
                        tenors=config.SPREAD_MONITOR_TENORS,
                    )
                    yield_data = {}
                    break

    if progress:
        progress("计算利差并生成前端数据", 78)
    cache = load_spread_history_cache()
    cached_dates = cache.setdefault("dates", {})
    rows = []
    daily_snapshots: dict[str, dict[str, dict]] = {}
    for bond in bonds:
        raw_code = str(bond["code"] or "").strip().upper()
        symbol = raw_code.split(".")[0]
        current_dt = dates.get("当前")
        current_yield = bond_yield_for(yield_data.get("当前", {}), raw_code)
        if current_yield is None:
            current_yield = cached_dates.get(current_dt, {}).get(symbol, {}).get("yield")
        if current_yield is None:
            continue
        curve_key = curve_for_bond(bond)
        spreads = {}
        current_curve_yield = None
        for label, dt in dates.items():
            cached_point = cached_dates.get(dt, {}).get(symbol, {})
            by_tenor = curve_series.get(curve_key, {})
            long_by_tenor = curve_series.get(config.LONG_END_CURVE, {})
            cy = interpolate_curve_with_long_end(by_tenor, long_by_tenor, bond["term"], dt)
            by = bond_yield_for(yield_data.get(label, {}), raw_code)
            if cy is None:
                cy = cached_point.get("curve_yield")
            if by is None:
                by = cached_point.get("yield")
            if label == "当前":
                current_curve_yield = cy
            spreads[label] = round((by - cy) * 100, 2) if by is not None and cy is not None else None
            if by is not None or cy is not None or spreads[label] is not None:
                daily_snapshots.setdefault(dt, {})[symbol] = {
                    "yield": by,
                    "curve_yield": cy,
                    "spread": spreads[label],
                }
        cur = spreads.get("当前")
        rows.append({
            "code": bond["code"],
            "raw_code": bond["raw_code"],
            "name": bond["name"],
            "implied_rating": bond.get("implied_rating") or "",
            "term": bond["term"],
            "is_holding": bond.get("is_holding", False),
            "curve_name": curve_key,
            "current_yield": round(current_yield, 4),
            "current_curve_yield": round(current_curve_yield, 4) if current_curve_yield is not None else None,
            "spread_current": cur,
            "spread_yesterday": spreads.get("昨日"),
            "spread_week": spreads.get("一周前"),
            "spread_month": spreads.get("一月前"),
            "spread_year_start": spreads.get("年初"),
            "change_daily": _diff(cur, spreads.get("昨日")),
            "change_weekly": _diff(cur, spreads.get("一周前")),
            "change_monthly": _diff(cur, spreads.get("一月前")),
            "change_ytd": _diff(cur, spreads.get("年初")),
        })
    for dt, snapshot in daily_snapshots.items():
        cached_dates.setdefault(dt, {}).update(snapshot)
    output = {
        "update_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "dates": dates,
        "total_bonds": len(rows),
        "holding_bonds": sum(1 for r in rows if r.get("is_holding")),
        "data": rows,
    }
    if not rows:
        raise RuntimeError(
            "利差监控生成结果为空，已保留现有数据文件；请检查中债估值表 TQ_QT_CBESTIMATE 最新日期"
        )
    save_spread_history_cache(cache)
    write_js(config.SPREAD_JS, output)
    return output


def _diff(a, b):
    return round(a - b, 2) if a is not None and b is not None else None


def generate_strategy_dashboard(progress=None) -> dict:
    return build_strategy_dashboard(progress=progress)


def generate_bond_picker_yields(progress=None) -> dict:
    if progress:
        progress("读取择券工具债券清单", 5)
    bonds = get_bond_picker_bonds()
    if not bonds:
        raise RuntimeError("未找到符合 BBB- 及以上且无担保人的择券工具债券")
    with connect() as conn:
        trade_date = latest_curve_date(conn)
        if progress:
            progress(f"查询择券工具最新中债估值 {dash_date(trade_date)}", 45)
        yields = fetch_cnbd_yields_by_symbol(conn, [b["code"] for b in bonds], trade_date)
    payload = save_bond_picker_yields_cache(trade_date, yields)
    if progress:
        progress(f"择券工具估值缓存完成 {len(payload['yields'])} 条", 90)
    return payload


def _strategy_html(payload: dict) -> str:
    data = json.dumps(payload, ensure_ascii=False)
    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>信用债策略仪表盘</title>
<script src="https://cdn.jsdelivr.net/npm/echarts@5/dist/echarts.min.js"></script>
<style>
*{{box-sizing:border-box}}body{{margin:0;background:#f6f7fb;color:#172033;font-family:-apple-system,BlinkMacSystemFont,"Segoe UI","Microsoft YaHei",sans-serif}}.main{{max-width:1180px;margin:0 auto;padding:22px}}h1{{font-size:22px;margin:0 0 6px}}.meta{{color:#64748b;font-size:13px;margin-bottom:18px}}.grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(360px,1fr));gap:14px}}.panel{{background:#fff;border:1px solid #e5e7eb;border-radius:8px;padding:14px}}.chart{{height:320px}}table{{width:100%;border-collapse:collapse;background:#fff;border:1px solid #e5e7eb}}th,td{{padding:9px;border-bottom:1px solid #eef2f7;text-align:right;font-size:13px}}th:first-child,td:first-child{{text-align:left}}
</style>
</head>
<body><main class="main">
<h1>信用债策略仪表盘</h1>
<div class="meta">数据日期: <span id="dt"></span> / 生成时间: <span id="gen"></span></div>
<div class="grid"><div class="panel"><div id="curveChart" class="chart"></div></div><div class="panel"><div id="benchChart" class="chart"></div></div></div>
<h2 style="font-size:16px;margin:18px 0 10px">最新曲线点</h2><div id="table"></div>
</main>
<script>
const DATA = {data};
document.getElementById('dt').textContent = DATA.latest_date;
document.getElementById('gen').textContent = DATA.generated_at;
const c1 = echarts.init(document.getElementById('curveChart'));
c1.setOption({{tooltip:{{trigger:'axis'}},legend:{{top:0}},grid:{{top:40,left:45,right:20,bottom:35}},xAxis:{{type:'category',data:[0,1,2,3,4,5,6,7,8,9,10].map(x=>x+'Y')}},yAxis:{{type:'value',name:'%'}},series:DATA.curves.map(c=>({{name:c.name,type:'line',smooth:true,data:[0,1,2,3,4,5,6,7,8,9,10].map(t=>{{const p=c.points.find(x=>x.tenor===t);return p?p.yield:null;}})}}))}});
const c2 = echarts.init(document.getElementById('benchChart'));
c2.setOption({{title:{{text:DATA.benchmark.length?(DATA.benchmark[0].name||DATA.benchmark_keyword):DATA.benchmark_keyword,left:0,textStyle:{{fontSize:13}}}},tooltip:{{trigger:'axis'}},grid:{{top:40,left:45,right:20,bottom:35}},xAxis:{{type:'category',data:DATA.benchmark.map(x=>String(x.date))}},yAxis:{{type:'value'}},series:[{{type:'line',smooth:true,data:DATA.benchmark.map(x=>x.close)}}]}});
let html='<table><thead><tr><th>曲线</th>'+[0,1,2,3,4,5,6,7,8,9,10].map(t=>'<th>'+t+'Y</th>').join('')+'</tr></thead><tbody>';
for (const c of DATA.curves) {{ html += '<tr><td>'+c.name+'</td>'+[0,1,2,3,4,5,6,7,8,9,10].map(t=>{{const p=c.points.find(x=>x.tenor===t);return '<td>'+(p?p.yield.toFixed(4):'-')+'</td>';}}).join('')+'</tr>'; }}
document.getElementById('table').innerHTML = html + '</tbody></table>';
window.addEventListener('resize',()=>{{c1.resize();c2.resize();}});
</script></body></html>"""


def run_all(progress=None, modules: list[str] | None = None) -> dict:
    selected = modules or ["bond_picker", "spread_monitor", "strategy_dashboard", "credit_std_dev", "institution_flow_rates"]
    selected = [m for m in selected if m in {"bond_picker", "spread_monitor", "strategy_dashboard", "credit_std_dev", "institution_flow_rates"}]
    if not selected:
        raise RuntimeError("未选择任何更新模块")
    settings = load_update_settings()
    with connect() as conn:
        term_target_date = latest_curve_date(conn)
    term_result = refresh_bond_terms(
        term_target_date,
        base_date=settings["excel_term_base_date"],
    )
    if progress:
        progress(
            "重算待偿期限："
            f"Excel 基准日 {term_result['base_date']} -> 最近交易日 {term_result['target_trade_date']}，"
            f"{term_result['updated']}/{term_result['total_bonds']} 条",
            2,
        )
    total = len(selected)
    results = {"bond_terms": term_result}
    for idx, module in enumerate(selected, start=1):
        base = int((idx - 1) / total * 100)
        end = int(idx / total * 100)

        def module_progress(message, percent=None, *, _base=base, _end=end):
            if not progress:
                return
            if percent is None:
                progress(message, None)
                return
            scaled = _base + int(max(0, min(100, int(percent))) / 100 * (_end - _base))
            progress(message, scaled)

        if module == "bond_picker":
            module_progress("开始更新择券工具估值", 1)
            results["bond_picker"] = generate_bond_picker_yields(progress=module_progress)
            module_progress("择券工具估值更新完成", 100)
        elif module == "spread_monitor":
            module_progress("开始更新利差监控", 1)
            results["spread_monitor"] = generate_spread_monitor(progress=module_progress)
            module_progress("利差监控更新完成", 100)
        elif module == "strategy_dashboard":
            module_progress("开始更新策略仪表盘", 1)
            results["strategy_dashboard"] = generate_strategy_dashboard(progress=module_progress)
            module_progress("策略仪表盘更新完成", 100)
        elif module == "credit_std_dev":
            if config.UPDATE_STD_DEV:
                module_progress("开始更新两倍标准差数据", 1)
                results["credit_std_dev"] = generate_std_dev_data(progress=module_progress)
                module_progress("两倍标准差数据更新完成", 100)
            else:
                module_progress("跳过两倍标准差数据", 100)
                results["credit_std_dev"] = "skipped"
        elif module == "institution_flow_rates":
            from .institution_flow_rates import update_rate_curves

            module_progress("开始更新机构行为国债/国开债/地方债曲线", 1)
            results["institution_flow_rates"] = update_rate_curves(progress=module_progress)
            module_progress("机构行为曲线更新完成", 100)
    return results
